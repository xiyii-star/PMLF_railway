"""
主评估脚本
运行 Baseline 和 SocketMatch 方法的评估，并比较结果
"""

import json
import logging
import argparse
from pathlib import Path
from datetime import datetime
import sys

# 添加项目根目录和src目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root / "src"))
sys.path.insert(0, str(Path(__file__).parent / "src"))

import openpyxl
from src.baseline_evaluator import BaselineEvaluator
from src.socketmatch_evaluator import SocketMatchEvaluator
from src.metrics_calculator import MetricsCalculator
from llm_config import create_llm_client

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('evaluation.log', encoding='utf-8')
    ]
)

logger = logging.getLogger(__name__)


def load_dataset_from_excel(excel_path: str) -> list:
    """
    从 Excel 文件加载数据集

    Args:
        excel_path: Excel 文件路径

    Returns:
        数据集列表
    """
    logger.info(f"从 Excel 加载数据集: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 获取列名（第一行）
    headers = [cell.value for cell in ws[1]]

    # 读取数据
    dataset = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, start=1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[header] = cell_value if cell_value is not None else ""

        dataset.append(row_data)

    logger.info(f"✅ 加载完成，共 {len(dataset)} 条数据")
    logger.info(f"   列名: {headers}")

    return dataset


def save_predictions_to_json(predictions: list, output_path: str):
    """
    保存预测结果到 JSON 文件

    Args:
        predictions: 预测结果列表
        output_path: 输出文件路径
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(predictions, f, ensure_ascii=False, indent=2)

    logger.info(f"预测结果已保存到: {output_path}")


def save_report_to_json(report: dict, output_path: str):
    """
    保存评估报告到 JSON 文件

    Args:
        report: 评估报告
        output_path: 输出文件路径
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    logger.info(f"评估报告已保存到: {output_path}")


def run_baseline_evaluation(
    dataset: list,
    llm_client,
    output_dir: Path
) -> tuple:
    """
    运行 Baseline 评估

    Args:
        dataset: 数据集
        llm_client: LLM 客户端
        output_dir: 输出目录

    Returns:
        (predictions, report)
    """
    logger.info("\n" + "="*80)
    logger.info("开始 Baseline 评估（Zero-shot LLM + Abstract Only）")
    logger.info("="*80)

    # 创建评估器
    prompts_dir = Path(__file__).parent.parent / "prompts"
    evaluator = BaselineEvaluator(llm_client, prompts_dir=str(prompts_dir))

    # 运行评估
    predictions, statistics = evaluator.evaluate_dataset(dataset)

    # 计算指标
    calculator = MetricsCalculator()
    report = calculator.generate_full_report(predictions, method_name="Baseline (Abstract Only)")

    # 打印报告
    calculator.print_report(report)

    # 保存结果
    save_predictions_to_json(predictions, str(output_dir / "baseline_predictions.json"))
    save_report_to_json(report, str(output_dir / "baseline_report.json"))

    return predictions, report


def run_socketmatch_evaluation(
    dataset: list,
    llm_client,
    output_dir: Path
) -> tuple:
    """
    运行 SocketMatch 评估

    Args:
        dataset: 数据集
        llm_client: LLM 客户端
        output_dir: 输出目录

    Returns:
        (predictions, report)
    """
    logger.info("\n" + "="*80)
    logger.info("开始 SocketMatch 评估（LLM + DeepPaper + SocketMatch Logic）")
    logger.info("="*80)

    # 创建评估器
    # 项目根目录: eval/citation_eval/run_evaluation.py -> parent -> parent -> parent
    prompts_dir = Path(__file__).parent.parent.parent / "prompts"
    evaluator = SocketMatchEvaluator(llm_client, prompts_dir=str(prompts_dir))

    # 运行评估
    predictions, statistics = evaluator.evaluate_dataset(dataset)

    # 计算指标
    calculator = MetricsCalculator()
    report = calculator.generate_full_report(predictions, method_name="SocketMatch (Deep Info + Logic)")

    # 打印报告
    calculator.print_report(report)

    # 保存结果
    save_predictions_to_json(predictions, str(output_dir / "socketmatch_predictions.json"))
    save_report_to_json(report, str(output_dir / "socketmatch_report.json"))

    return predictions, report


def compare_methods(
    baseline_report: dict,
    socketmatch_report: dict,
    output_dir: Path
):
    """
    比较两种方法

    Args:
        baseline_report: Baseline 评估报告
        socketmatch_report: SocketMatch 评估报告
        output_dir: 输出目录
    """
    logger.info("\n" + "="*80)
    logger.info("方法比较")
    logger.info("="*80)

    calculator = MetricsCalculator()

    # 生成比较结果
    comparison = calculator.compare_methods([baseline_report, socketmatch_report])

    # 打印比较结果
    calculator.print_comparison(comparison)

    # 保存比较结果
    save_report_to_json(comparison, str(output_dir / "comparison.json"))

    # 生成性能提升报告
    baseline_acc = baseline_report['overall_metrics']['accuracy']
    socketmatch_acc = socketmatch_report['overall_metrics']['accuracy']
    baseline_f1 = baseline_report['overall_metrics']['macro_f1']
    socketmatch_f1 = socketmatch_report['overall_metrics']['macro_f1']

    improvement = {
        'accuracy_improvement': socketmatch_acc - baseline_acc,
        'accuracy_improvement_percent': ((socketmatch_acc - baseline_acc) / baseline_acc * 100) if baseline_acc > 0 else 0,
        'macro_f1_improvement': socketmatch_f1 - baseline_f1,
        'macro_f1_improvement_percent': ((socketmatch_f1 - baseline_f1) / baseline_f1 * 100) if baseline_f1 > 0 else 0
    }

    logger.info("\n【性能提升】")
    logger.info(f"  Accuracy 提升: {improvement['accuracy_improvement']:+.4f} ({improvement['accuracy_improvement_percent']:+.2f}%)")
    logger.info(f"  Macro F1 提升: {improvement['macro_f1_improvement']:+.4f} ({improvement['macro_f1_improvement_percent']:+.2f}%)")

    save_report_to_json(improvement, str(output_dir / "improvement.json"))


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='引用关系分类评估')
    parser.add_argument(
        '--data',
        type=str,
        default='eval/citation_eval/data/golden_citation_dataset.xlsx',
        help='数据集路径'
    )
    parser.add_argument(
        '--config',
        type=str,
        default='config/config.yaml',
        help='LLM 配置文件路径'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='eval/citation_eval/results',
        help='输出目录'
    )
    parser.add_argument(
        '--method',
        type=str,
        choices=['baseline', 'socketmatch', 'both'],
        default='both',
        help='运行的方法: baseline, socketmatch, both'
    )
    parser.add_argument(
        '--sample-size',
        type=int,
        default=None,
        help='采样大小（用于快速测试，None表示使用全部数据）'
    )

    args = parser.parse_args()

    # 创建输出目录
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 添加时间戳子目录
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_output_dir = output_dir / timestamp
    run_output_dir.mkdir(parents=True, exist_ok=True)

    logger.info("\n" + "="*80)
    logger.info("引用关系分类评估实验")
    logger.info("="*80)
    logger.info(f"数据集: {args.data}")
    logger.info(f"配置文件: {args.config}")
    logger.info(f"输出目录: {run_output_dir}")
    logger.info(f"运行方法: {args.method}")
    if args.sample_size:
        logger.info(f"采样大小: {args.sample_size}")
    logger.info("="*80)

    # 加载数据集
    dataset = load_dataset_from_excel(args.data)

    # 采样（如果指定）
    if args.sample_size and args.sample_size < len(dataset):
        import random
        random.seed(42)
        dataset = random.sample(dataset, args.sample_size)
        logger.info(f"采样 {args.sample_size} 条数据用于测试")

    # 创建 LLM 客户端
    logger.info(f"\n从配置文件加载 LLM 客户端: {args.config}")
    llm_client = create_llm_client(args.config)

    # 运行评估
    baseline_report = None
    socketmatch_report = None

    if args.method in ['baseline', 'both']:
        _, baseline_report = run_baseline_evaluation(dataset, llm_client, run_output_dir)

    if args.method in ['socketmatch', 'both']:
        _, socketmatch_report = run_socketmatch_evaluation(dataset, llm_client, run_output_dir)

    # 比较方法（如果两种方法都运行了）
    if args.method == 'both' and baseline_report and socketmatch_report:
        compare_methods(baseline_report, socketmatch_report, run_output_dir)

    logger.info("\n" + "="*80)
    logger.info("✅ 评估实验完成")
    logger.info(f"结果已保存到: {run_output_dir}")
    logger.info("="*80)


if __name__ == "__main__":
    main()
